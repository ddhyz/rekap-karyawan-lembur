# app_lembur.py
import streamlit as st
import pandas as pd
from datetime import time
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

# Konfigurasi halaman (judul tab browser)
st.set_page_config(page_title="Rekap Lembur PT. Quantum")

# Judul aplikasi
st.title("📊 Rekap Lembur PT. Quantum")

# Upload file absensi
uploaded_file = st.file_uploader("📂 Upload File Absensi Excel", type=["xlsx", "xls"])

if uploaded_file:
    # Baca data absensi
    df = pd.read_excel(uploaded_file)

    # Hapus kolom yang tidak perlu (jika ada)
    drop_cols = [
    "Lokasi ID", "ID Number", "VerifyCode", "CardNo", "Jam.",
    "No.PIN", "Kode Verifikasi", "No.Kartu"
]
    df = df.drop(columns=[c for c in drop_cols if c in df.columns], errors="ignore")

    # Pastikan ada kolom 'Tgl/Waktu'
    if "Tgl/Waktu" in df.columns:
        # Konversi ke datetime
        df["Tgl/Waktu"] = pd.to_datetime(df["Tgl/Waktu"], errors="coerce")

        # Ambil jam saja
        df["Jam"] = df["Tgl/Waktu"].dt.time

        # Filter data lembur (pulang > 18:05)
        lembur = df[df["Jam"] > time(18, 5)]

        # Filter data pulang lebih awal (17:40–18:04)
        pulang_awal = df[(df["Jam"] >= time(17, 40)) & (df["Jam"] <= time(18, 4))]

        # Tampilkan tabel di aplikasi
        st.write("### 👨‍💻 Karyawan Lembur (Pulang > 18:05)")
        st.dataframe(lembur)

        st.write("### 🏃 Karyawan Pulang Lebih Awal (17:40–18:04)")
        st.dataframe(pulang_awal)

        # Fungsi export ke Excel dengan border
        def to_excel(lembur, pulang_awal):
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                lembur.to_excel(writer, sheet_name="Lembur", index=False)
                pulang_awal.to_excel(writer, sheet_name="Pulang_Awal", index=False)

            # Buka workbook lagi untuk kasih border
            output.seek(0)
            wb = load_workbook(output)
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)

            for sheet_name in ["Lembur", "Pulang_Awal"]:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                        min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = border

            # Simpan ulang
            final_output = BytesIO()
            wb.save(final_output)
            return final_output.getvalue()

        # Buat file Excel untuk diunduh
        excel_data = to_excel(lembur, pulang_awal)

        st.download_button(
            label="📥 Download Rekap Excel",
            data=excel_data,
            file_name="Rekap_Lembur.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    else:
        st.error("❌ Kolom 'Tgl/Waktu' tidak ditemukan di file absensi!")

else:
    st.info("Silakan upload file absensi untuk diproses.")


